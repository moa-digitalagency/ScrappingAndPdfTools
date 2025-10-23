"""
PdfTools
MOA Digital Agency LLC
Service d'extraction de texte depuis des PDFs
"""

import os
import logging
import csv
from pypdf import PdfReader
from typing import Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

logger = logging.getLogger(__name__)

class PdfTextExtractor:
    """Service pour extraire le texte de fichiers PDF"""
    
    @staticmethod
    def extract_text_from_pdf(pdf_path: str) -> Dict:
        """
        Extrait le texte d'un fichier PDF
        
        Args:
            pdf_path: Chemin vers le fichier PDF
            
        Returns:
            Dict avec le texte extrait, nombre de pages, et métadonnées
        """
        try:
            reader = PdfReader(pdf_path)
            
            # Extraire le texte de toutes les pages
            full_text = ""
            page_texts = []
            
            for i, page in enumerate(reader.pages):
                page_text = page.extract_text()
                # PyPDF peut retourner None pour les pages sans texte (images, etc.)
                if page_text is None:
                    page_text = ""
                
                page_texts.append({
                    'page_number': i + 1,
                    'text': page_text,
                    'char_count': len(page_text)
                })
                full_text += page_text + "\n\n"
            
            # Récupérer les métadonnées
            metadata = {}
            if reader.metadata:
                metadata = {
                    'title': reader.metadata.get('/Title', ''),
                    'author': reader.metadata.get('/Author', ''),
                    'subject': reader.metadata.get('/Subject', ''),
                    'creator': reader.metadata.get('/Creator', ''),
                    'producer': reader.metadata.get('/Producer', ''),
                    'creation_date': reader.metadata.get('/CreationDate', ''),
                }
            
            # Nettoyer le texte et calculer les statistiques réelles
            cleaned_text = full_text.strip()
            
            return {
                'success': True,
                'text': cleaned_text,
                'page_count': len(reader.pages),
                'pages': page_texts,
                'metadata': metadata,
                'total_chars': len(cleaned_text),
                'total_words': len(cleaned_text.split()) if cleaned_text else 0
            }
            
        except Exception as e:
            logger.error(f"Erreur lors de l'extraction de texte de {pdf_path}: {e}")
            return {
                'success': False,
                'error': str(e)
            }
    
    @staticmethod
    def extract_text_from_multiple_pdfs(pdf_paths: List[str]) -> List[Dict]:
        """
        Extrait le texte de plusieurs fichiers PDF
        
        Args:
            pdf_paths: Liste des chemins vers les fichiers PDF
            
        Returns:
            Liste de dictionnaires avec les résultats d'extraction
        """
        results = []
        
        for pdf_path in pdf_paths:
            result = PdfTextExtractor.extract_text_from_pdf(pdf_path)
            result['file_path'] = pdf_path
            result['file_name'] = os.path.basename(pdf_path)
            results.append(result)
        
        return results
    
    @staticmethod
    def save_text_to_file(text: str, output_path: str) -> bool:
        """
        Sauvegarde le texte extrait dans un fichier
        
        Args:
            text: Texte à sauvegarder
            output_path: Chemin du fichier de sortie
            
        Returns:
            True si réussi, False sinon
        """
        try:
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(text)
            return True
        except Exception as e:
            logger.error(f"Erreur lors de la sauvegarde du texte: {e}")
            return False
    
    @staticmethod
    def export_to_excel(results: List[Dict], output_path: str) -> bool:
        """
        Exporte les résultats d'extraction vers Excel
        
        Args:
            results: Liste des résultats d'extraction
            output_path: Chemin du fichier Excel de sortie
            
        Returns:
            True si réussi, False sinon
        """
        try:
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            wb = Workbook()
            
            # Feuille de résumé
            ws_summary = wb.active
            ws_summary.title = "Résumé"
            
            # En-têtes pour le résumé
            headers_summary = ['Fichier', 'Pages', 'Caractères', 'Mots', 'Statut']
            ws_summary.append(headers_summary)
            
            # Style pour les en-têtes
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            for col_num, header in enumerate(headers_summary, 1):
                cell = ws_summary.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Données du résumé
            for result in results:
                if result.get('success'):
                    ws_summary.append([
                        result.get('file_name', 'Inconnu'),
                        result.get('page_count', 0),
                        result.get('total_chars', 0),
                        result.get('total_words', 0),
                        'Réussi'
                    ])
                else:
                    ws_summary.append([
                        result.get('file_name', 'Inconnu'),
                        0,
                        0,
                        0,
                        f"Échec: {result.get('error', 'Erreur inconnue')}"
                    ])
            
            # Ajuster la largeur des colonnes
            for col in ws_summary.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_summary.column_dimensions[column].width = adjusted_width
            
            # Feuille de texte complet
            ws_text = wb.create_sheet(title="Texte extrait")
            ws_text.append(['Fichier', 'Texte'])
            
            # Style pour les en-têtes
            for col_num in range(1, 3):
                cell = ws_text.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            for result in results:
                if result.get('success'):
                    ws_text.append([
                        result.get('file_name', 'Inconnu'),
                        result.get('text', '')
                    ])
            
            # Ajuster la largeur des colonnes
            ws_text.column_dimensions['A'].width = 30
            ws_text.column_dimensions['B'].width = 100
            
            wb.save(output_path)
            return True
            
        except Exception as e:
            logger.error(f"Erreur lors de l'export Excel: {e}")
            return False
    
    @staticmethod
    def export_to_csv(results: List[Dict], output_path: str) -> bool:
        """
        Exporte les résultats d'extraction vers CSV
        
        Args:
            results: Liste des résultats d'extraction
            output_path: Chemin du fichier CSV de sortie
            
        Returns:
            True si réussi, False sinon
        """
        try:
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            with open(output_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                fieldnames = ['Fichier', 'Pages', 'Caractères', 'Mots', 'Statut', 'Texte']
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                
                writer.writeheader()
                
                for result in results:
                    if result.get('success'):
                        writer.writerow({
                            'Fichier': result.get('file_name', 'Inconnu'),
                            'Pages': result.get('page_count', 0),
                            'Caractères': result.get('total_chars', 0),
                            'Mots': result.get('total_words', 0),
                            'Statut': 'Réussi',
                            'Texte': result.get('text', '')
                        })
                    else:
                        writer.writerow({
                            'Fichier': result.get('file_name', 'Inconnu'),
                            'Pages': 0,
                            'Caractères': 0,
                            'Mots': 0,
                            'Statut': f"Échec: {result.get('error', 'Erreur inconnue')}",
                            'Texte': ''
                        })
            
            return True
            
        except Exception as e:
            logger.error(f"Erreur lors de l'export CSV: {e}")
            return False
