import os
import zipfile
import uuid
from datetime import datetime
from pypdf import PdfWriter, PdfReader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def merge_pdfs_from_zip(zip_path, temp_folder):
    temp_dir = os.path.join(temp_folder, str(uuid.uuid4()))
    os.makedirs(temp_dir, exist_ok=True)
    
    extract_dir = os.path.join(temp_dir, 'extracted')
    os.makedirs(extract_dir, exist_ok=True)
    
    try:
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(extract_dir)
    except Exception as e:
        return {'success': False, 'error': f'Erreur lors de l\'extraction du ZIP: {str(e)}'}
    
    pdf_files = []
    for root, dirs, files in os.walk(extract_dir):
        for file in files:
            if file.lower().endswith('.pdf'):
                pdf_files.append(os.path.join(root, file))
    
    pdf_files.sort()
    
    if not pdf_files:
        return {'success': False, 'error': 'Aucun fichier PDF trouvé dans le ZIP'}
    
    merger = PdfWriter()
    pdf_metadata = []
    total_pages = 0
    
    for idx, pdf_file in enumerate(pdf_files, 1):
        try:
            reader = PdfReader(pdf_file)
            num_pages = len(reader.pages)
            total_pages += num_pages
            
            for page in reader.pages:
                merger.add_page(page)
            
            pdf_metadata.append({
                'index': idx,
                'filename': os.path.basename(pdf_file),
                'pages': num_pages,
                'path': pdf_file
            })
            
        except Exception as e:
            pdf_metadata.append({
                'index': idx,
                'filename': os.path.basename(pdf_file),
                'pages': 0,
                'error': str(e)
            })
    
    unique_id = str(uuid.uuid4())[:8]
    pdf_filename = f'merged_{unique_id}.pdf'
    pdf_path = os.path.join(temp_folder, pdf_filename)
    
    with open(pdf_path, 'wb') as output_pdf:
        merger.write(output_pdf)
    
    merger.close()
    
    excel_filename = f'manifest_{unique_id}.xlsx'
    excel_path = os.path.join(temp_folder, excel_filename)
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Manifest PDFs'
    
    headers = ['#', 'Nom du fichier', 'Nombre de pages', 'Statut']
    ws.append(headers)
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for item in pdf_metadata:
        status = 'Erreur' if 'error' in item else 'OK'
        ws.append([
            item['index'],
            item['filename'],
            item['pages'],
            status
        ])
    
    ws.append([])
    ws.append(['Total de PDFs:', len(pdf_files)])
    ws.append(['Total de pages:', total_pages])
    
    for row in ws.iter_rows(min_row=len(pdf_metadata) + 2, max_row=len(pdf_metadata) + 4):
        for cell in row:
            cell.font = Font(bold=True)
    
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    
    wb.save(excel_path)
    
    for root, dirs, files in os.walk(temp_dir, topdown=False):
        for file in files:
            os.remove(os.path.join(root, file))
        for dir in dirs:
            try:
                os.rmdir(os.path.join(root, dir))
            except:
                pass
    try:
        os.rmdir(temp_dir)
    except:
        pass
    
    return {
        'success': True,
        'pdf_path': pdf_path,
        'excel_path': excel_path,
        'pdf_filename': pdf_filename,
        'excel_filename': excel_filename,
        'total_pdfs': len(pdf_files),
        'total_pages': total_pages
    }
