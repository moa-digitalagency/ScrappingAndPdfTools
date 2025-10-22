"""
PdfTools
MOA Digital Agency LLC
Par : Aisance KALONJI
Mail : moa@myoneart.com
www.myoneart.com
"""

import os
import requests
import logging
from pypdf import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def extract_text_from_pdf(pdf_path):
    """Extrait le texte d'un fichier PDF"""
    try:
        reader = PdfReader(pdf_path)
        text = ""
        for page in reader.pages:
            text += page.extract_text() + "\n"
        return text.strip()
    except Exception as e:
        logger.error(f"Erreur extraction texte de {pdf_path}: {str(e)}")
        return None

def analyze_pdf_with_ai(pdf_text, filename, api_key):
    """Analyse un PDF avec OpenRouter API pour extraire la structure intelligente"""
    try:
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            "HTTP-Referer": "https://pdf-tools.replit.app",
            "X-Title": "PDF Tools Analyzer"
        }
        
        prompt = f"""Analysez ce document PDF et extrayez les informations structurées suivantes au format JSON:

Nom du fichier: {filename}

Contenu du PDF:
{pdf_text[:15000]}

Retournez un JSON avec cette structure:
{{
  "titre": "titre du document",
  "type_document": "type (rapport, facture, contrat, etc.)",
  "date": "date si trouvée",
  "entites": ["liste des entités/organisations mentionnées"],
  "mots_cles": ["liste des mots-clés importants"],
  "resume": "résumé en 2-3 phrases",
  "champs_personnalises": {{"champ1": "valeur1", "champ2": "valeur2"}}
}}

Si certaines informations ne sont pas disponibles, utilisez "N/A"."""

        data = {
            "model": "meta-llama/llama-3.1-8b-instruct:free",
            "messages": [
                {
                    "role": "user",
                    "content": prompt
                }
            ],
            "temperature": 0.3,
            "max_tokens": 2000
        }
        
        response = requests.post(
            "https://openrouter.ai/api/v1/chat/completions",
            headers=headers,
            json=data,
            timeout=60
        )
        response.raise_for_status()
        
        result = response.json()
        ai_response = result['choices'][0]['message']['content']
        
        import json
        ai_response_clean = ai_response.strip()
        if ai_response_clean.startswith('```json'):
            ai_response_clean = ai_response_clean[7:]
        if ai_response_clean.startswith('```'):
            ai_response_clean = ai_response_clean[3:]
        if ai_response_clean.endswith('```'):
            ai_response_clean = ai_response_clean[:-3]
        
        analysis = json.loads(ai_response_clean.strip())
        return analysis
        
    except Exception as e:
        logger.error(f"Erreur analyse AI pour {filename}: {str(e)}")
        return {
            "titre": filename,
            "type_document": "N/A",
            "date": "N/A",
            "entites": [],
            "mots_cles": [],
            "resume": f"Erreur d'analyse: {str(e)}",
            "champs_personnalises": {}
        }

def analyze_pdfs_and_create_database(pdf_files, temp_folder):
    """Analyse plusieurs PDFs et crée une base de données Excel exportable"""
    
    api_key = os.environ.get('OPENROUTER_API_KEY')
    if not api_key:
        return {
            'success': False,
            'error': 'Clé API OpenRouter non configurée. Veuillez configurer OPENROUTER_API_KEY dans les secrets.'
        }
    
    analyses = []
    total_files = len(pdf_files)
    
    logger.info(f"Début de l'analyse intelligente de {total_files} PDFs")
    
    for idx, pdf_path in enumerate(pdf_files, 1):
        filename = os.path.basename(pdf_path)
        logger.info(f"Analyse {idx}/{total_files}: {filename}")
        
        text = extract_text_from_pdf(pdf_path)
        
        if text and len(text) > 50:
            analysis = analyze_pdf_with_ai(text, filename, api_key)
            analysis['fichier'] = filename
            analysis['longueur_texte'] = len(text)
            analysis['pages'] = len(PdfReader(pdf_path).pages)
            analyses.append(analysis)
        else:
            analyses.append({
                'fichier': filename,
                'titre': filename,
                'type_document': 'N/A',
                'date': 'N/A',
                'entites': [],
                'mots_cles': [],
                'resume': 'Texte insuffisant pour analyse',
                'pages': len(PdfReader(pdf_path).pages) if os.path.exists(pdf_path) else 0,
                'longueur_texte': len(text) if text else 0,
                'champs_personnalises': {}
            })
    
    import uuid
    unique_id = str(uuid.uuid4())[:8]
    excel_filename = f'analyse_intelligente_{unique_id}.xlsx'
    excel_path = os.path.join(temp_folder, excel_filename)
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Analyse Intelligente PDFs'
    
    headers = ['Fichier', 'Titre', 'Type', 'Date', 'Pages', 'Longueur Texte', 'Entités', 'Mots-clés', 'Résumé']
    ws.append(headers)
    
    header_fill = PatternFill(start_color='1F4788', end_color='1F4788', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for analysis in analyses:
        ws.append([
            analysis.get('fichier', 'N/A'),
            analysis.get('titre', 'N/A'),
            analysis.get('type_document', 'N/A'),
            analysis.get('date', 'N/A'),
            analysis.get('pages', 0),
            analysis.get('longueur_texte', 0),
            ', '.join(analysis.get('entites', [])),
            ', '.join(analysis.get('mots_cles', [])),
            analysis.get('resume', 'N/A')
        ])
    
    for row in ws.iter_rows(min_row=2, max_row=len(analyses) + 1):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 40
    ws.column_dimensions['H'].width = 40
    ws.column_dimensions['I'].width = 60
    
    if analyses:
        ws2 = wb.create_sheet('Champs Personnalisés')
        custom_headers = ['Fichier'] + list(set([key for a in analyses for key in a.get('champs_personnalises', {}).keys()]))
        ws2.append(custom_headers)
        
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for analysis in analyses:
            row_data = [analysis.get('fichier', 'N/A')]
            custom_fields = analysis.get('champs_personnalises', {})
            for header in custom_headers[1:]:
                row_data.append(custom_fields.get(header, 'N/A'))
            ws2.append(row_data)
    
    wb.save(excel_path)
    
    logger.info(f"Analyse terminée. Base de données Excel créée: {excel_filename}")
    
    return {
        'success': True,
        'excel_path': excel_path,
        'excel_filename': excel_filename,
        'total_analyzed': len(analyses),
        'analyses': analyses
    }
