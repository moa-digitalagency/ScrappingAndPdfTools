"""
PdfTools
MOA Digital Agency LLC
Par : Aisance KALONJI
Mail : moa@myoneart.com
www.myoneart.com
"""

import os
import requests
import logging
import json
import uuid
from pypdf import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from concurrent.futures import ThreadPoolExecutor, as_completed
import zipfile
import csv
import io
from config import Config

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def extract_text_from_pdf(pdf_path):
    """Extrait le texte d'un fichier PDF"""
    try:
        reader = PdfReader(pdf_path)
        text = ""
        for page in reader.pages:
            text += page.extract_text() + "\n"
        return text.strip()
    except Exception as e:
        logger.error(f"Erreur extraction texte de {pdf_path}: {str(e)}")
        return None

def extract_jurisprudence_data_with_ai(pdf_text, filename, api_key, pdf_path=None, num_pages=0):
    """Extrait les données de jurisprudence avec OpenRouter API"""
    try:
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            "HTTP-Referer": "https://pdf-tools.replit.app",
            "X-Title": "PDF Tools Jurisprudence Extractor"
        }
        
        prompt = f"""Tu es un expert en extraction de données juridiques. Analyse ce document de jurisprudence et extrait TOUTES les informations suivantes au format JSON strict.

Nom du fichier: {filename}

Contenu du document:
{pdf_text[:20000]}

Retourne UN SEUL objet JSON avec cette structure EXACTE (utilise "N/A" si l'information n'est pas disponible):

{{
  "ref": "numéro de référence",
  "titre": "titre complet de la décision",
  "juridiction": "nom de la juridiction",
  "pays_ville": "pays et ville",
  "numero_decision": "numéro de la décision",
  "date_decision": "date de la décision",
  "numero_dossier": "numéro de dossier",
  "type_decision": "type (Arrêt, Jugement, etc.)",
  "chambre": "chambre (Commerciale, Civile, etc.)",
  "theme": "thème principal",
  "mots_cles": "tous les mots-clés séparés par des virgules",
  "base_legale_articles": "articles de loi cités",
  "base_legale_lois": "lois citées",
  "resume_francais": "résumé complet en français",
  "resume_arabe": "résumé en arabe si disponible",
  "texte_integral_debut": "les 1000 premiers caractères du texte intégral",
  "source": "source de publication"
}}

IMPORTANT: Retourne UNIQUEMENT le JSON, sans texte avant ou après, sans ```json```, juste l'objet JSON pur."""

        data = {
            "model": "meta-llama/llama-3.1-8b-instruct:free",
            "messages": [
                {
                    "role": "user",
                    "content": prompt
                }
            ],
            "temperature": 0.1,
            "max_tokens": 3000
        }
        
        response = requests.post(
            "https://openrouter.ai/api/v1/chat/completions",
            headers=headers,
            json=data,
            timeout=90
        )
        
        if response.status_code == 404:
            error_detail = response.text[:500] if len(response.text) > 500 else response.text
            raise Exception(f"Erreur 404 de l'API OpenRouter. Cela peut signifier: (1) Clé API invalide, (2) Modèle non disponible, ou (3) Endpoint incorrect. Détails: {error_detail}")
        elif response.status_code == 429:
            raise Exception("Limite de taux atteinte. Les modèles gratuits sont limités à 50 requêtes/jour (ou 1000/jour avec 10$ de crédits). Veuillez attendre ou acheter des crédits sur https://openrouter.ai")
        elif response.status_code == 401:
            raise Exception("Clé API invalide ou expirée. Veuillez vérifier votre clé API OpenRouter.")
        
        response.raise_for_status()
        
        # Vérifier si la réponse est du JSON valide
        try:
            result = response.json()
        except json.JSONDecodeError:
            # La réponse n'est pas du JSON (probablement du HTML)
            error_preview = response.text[:500] if len(response.text) > 500 else response.text
            logger.error(f"Réponse non-JSON de l'API: {error_preview}")
            raise Exception(f"L'API a renvoyé une réponse invalide (HTML au lieu de JSON). Cela peut indiquer un problème de quota ou de limite de taux. Réponse: {error_preview}")
        
        # Vérifier que la structure de la réponse est correcte
        if 'choices' not in result or not result['choices']:
            error_msg = result.get('error', {}).get('message', 'Structure de réponse invalide')
            raise Exception(f"Erreur de l'API OpenRouter: {error_msg}")
        
        ai_response = result['choices'][0]['message']['content']
        
        # Nettoyer la réponse
        ai_response_clean = ai_response.strip()
        if ai_response_clean.startswith('```json'):
            ai_response_clean = ai_response_clean[7:]
        if ai_response_clean.startswith('```'):
            ai_response_clean = ai_response_clean[3:]
        if ai_response_clean.endswith('```'):
            ai_response_clean = ai_response_clean[:-3]
        
        jurisprudence_data = json.loads(ai_response_clean.strip())
        
        # Ajouter le nom du fichier
        jurisprudence_data['fichier'] = filename
        
        return jurisprudence_data
        
    except Exception as e:
        logger.error(f"Erreur extraction jurisprudence pour {filename}: {str(e)}")
        return {
            "fichier": filename,
            "ref": "N/A",
            "titre": "N/A",
            "juridiction": "N/A",
            "pays_ville": "N/A",
            "numero_decision": "N/A",
            "date_decision": "N/A",
            "numero_dossier": "N/A",
            "type_decision": "N/A",
            "chambre": "N/A",
            "theme": "N/A",
            "mots_cles": "N/A",
            "base_legale_articles": "N/A",
            "base_legale_lois": "N/A",
            "resume_francais": f"Erreur d'extraction: {str(e)}",
            "resume_arabe": "N/A",
            "texte_integral_debut": "N/A",
            "source": "N/A"
        }

def process_single_pdf_jurisprudence(pdf_path, filename, api_key):
    """Traite un seul PDF de jurisprudence"""
    logger.info(f"Extraction jurisprudence: {filename}")
    
    text = extract_text_from_pdf(pdf_path)
    if not text or len(text) < 50:
        return {
            'success': False,
            'filename': filename,
            'error': 'Texte insuffisant pour extraction'
        }
    
    # Obtenir le nombre de pages
    try:
        num_pages = len(PdfReader(pdf_path).pages)
    except Exception:
        num_pages = 0
    
    jurisprudence_data = extract_jurisprudence_data_with_ai(text, filename, api_key, pdf_path, num_pages)
    return {
        'success': True,
        'filename': filename,
        'data': jurisprudence_data
    }

def create_jurisprudence_excel(jurisprudence_list, temp_folder):
    """Crée un fichier Excel avec les données de jurisprudence"""
    unique_id = str(uuid.uuid4())[:8]
    excel_filename = f'jurisprudence_database_{unique_id}.xlsx'
    excel_path = os.path.join(temp_folder, excel_filename)
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Base Jurisprudence'
    
    # En-têtes
    headers = [
        'Fichier', 'Ref', 'Titre', 'Juridiction', 'Pays/Ville', 
        'N° Décision', 'Date Décision', 'N° Dossier', 'Type Décision',
        'Chambre', 'Thème', 'Mots-clés', 'Articles', 'Lois',
        'Résumé Français', 'Résumé Arabe', 'Extrait Texte Intégral (Arabe)', 'Source'
    ]
    ws.append(headers)
    
    # Style des en-têtes
    header_fill = PatternFill(start_color='1F4788', end_color='1F4788', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Données
    for item in jurisprudence_list:
        data = item.get('data', {})
        ws.append([
            data.get('fichier', 'N/A'),
            data.get('ref', 'N/A'),
            data.get('titre', 'N/A'),
            data.get('juridiction', 'N/A'),
            data.get('pays_ville', 'N/A'),
            data.get('numero_decision', 'N/A'),
            data.get('date_decision', 'N/A'),
            data.get('numero_dossier', 'N/A'),
            data.get('type_decision', 'N/A'),
            data.get('chambre', 'N/A'),
            data.get('theme', 'N/A'),
            data.get('mots_cles', 'N/A'),
            data.get('base_legale_articles', 'N/A'),
            data.get('base_legale_lois', 'N/A'),
            data.get('resume_francais', 'N/A'),
            data.get('resume_arabe', 'N/A'),
            data.get('texte_integral_debut', 'N/A'),
            data.get('source', 'N/A')
        ])
    
    # Style des données
    for row in ws.iter_rows(min_row=2, max_row=len(jurisprudence_list) + 1):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = border
    
    # Largeurs de colonnes
    ws.column_dimensions['A'].width = 30  # Fichier
    ws.column_dimensions['B'].width = 12  # Ref
    ws.column_dimensions['C'].width = 50  # Titre
    ws.column_dimensions['D'].width = 25  # Juridiction
    ws.column_dimensions['E'].width = 20  # Pays/Ville
    ws.column_dimensions['F'].width = 15  # N° Décision
    ws.column_dimensions['G'].width = 15  # Date
    ws.column_dimensions['H'].width = 20  # N° Dossier
    ws.column_dimensions['I'].width = 15  # Type
    ws.column_dimensions['J'].width = 15  # Chambre
    ws.column_dimensions['K'].width = 30  # Thème
    ws.column_dimensions['L'].width = 50  # Mots-clés
    ws.column_dimensions['M'].width = 30  # Articles
    ws.column_dimensions['N'].width = 40  # Lois
    ws.column_dimensions['O'].width = 70  # Résumé FR
    ws.column_dimensions['P'].width = 70  # Résumé AR
    ws.column_dimensions['Q'].width = 50  # Extrait
    ws.column_dimensions['R'].width = 20  # Source
    ws.column_dimensions['S'].width = 10  # Pages
    ws.column_dimensions['T'].width = 12  # Taille
    
    wb.save(excel_path)
    logger.info(f"Base de données Excel créée: {excel_filename}")
    
    return excel_path, excel_filename

def create_jurisprudence_csv(jurisprudence_list, temp_folder):
    """Crée un fichier CSV avec les données de jurisprudence"""
    unique_id = str(uuid.uuid4())[:8]
    csv_filename = f'jurisprudence_database_{unique_id}.csv'
    csv_path = os.path.join(temp_folder, csv_filename)
    
    headers = [
        'Fichier', 'Ref', 'Titre', 'Juridiction', 'Pays/Ville', 
        'N° Décision', 'Date Décision', 'N° Dossier', 'Type Décision',
        'Chambre', 'Thème', 'Mots-clés', 'Articles', 'Lois',
        'Résumé Français', 'Résumé Arabe', 'Extrait Texte', 'Source',
        'Pages', 'Taille Texte'
    ]
    
    with open(csv_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(headers)
        
        for item in jurisprudence_list:
            data = item.get('data', {})
            writer.writerow([
                data.get('fichier', 'N/A'),
                data.get('ref', 'N/A'),
                data.get('titre', 'N/A'),
                data.get('juridiction', 'N/A'),
                data.get('pays_ville', 'N/A'),
                data.get('numero_decision', 'N/A'),
                data.get('date_decision', 'N/A'),
                data.get('numero_dossier', 'N/A'),
                data.get('type_decision', 'N/A'),
                data.get('chambre', 'N/A'),
                data.get('theme', 'N/A'),
                data.get('mots_cles', 'N/A'),
                data.get('base_legale_articles', 'N/A'),
                data.get('base_legale_lois', 'N/A'),
                data.get('resume_francais', 'N/A'),
                data.get('resume_arabe', 'N/A'),
                data.get('texte_integral_debut', 'N/A'),
                data.get('source', 'N/A')
            ])
    
    logger.info(f"Base de données CSV créée: {csv_filename}")
    return csv_path, csv_filename

def extract_jurisprudence_from_zip(zip_path, temp_folder, output_format='excel'):
    """Extrait la jurisprudence depuis un ZIP de PDFs"""
    # Charger la clé API depuis .env (VPS) ou secrets Replit
    api_key = Config.OPENROUTER_API_KEY
    if not api_key:
        return {
            'success': False,
            'error': 'Clé API OpenRouter non configurée. Veuillez configurer OPENROUTER_API_KEY dans les secrets ou le fichier .env.'
        }
    
    try:
        # Extraire le ZIP
        extract_dir = os.path.join(temp_folder, f'jurisprudence_extract_{uuid.uuid4().hex[:8]}')
        os.makedirs(extract_dir, exist_ok=True)
        
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(extract_dir)
        
        # Trouver tous les PDFs
        pdf_files = []
        for root, dirs, files in os.walk(extract_dir):
            for file in files:
                if file.lower().endswith('.pdf'):
                    pdf_files.append(os.path.join(root, file))
        
        if not pdf_files:
            return {
                'success': False,
                'error': 'Aucun fichier PDF trouvé dans le ZIP'
            }
        
        logger.info(f"Traitement de {len(pdf_files)} fichiers PDF de jurisprudence")
        
        # Traiter les PDFs en parallèle
        jurisprudence_list = []
        with ThreadPoolExecutor(max_workers=3) as executor:
            futures = {
                executor.submit(
                    process_single_pdf_jurisprudence,
                    pdf_path,
                    os.path.basename(pdf_path),
                    api_key
                ): pdf_path for pdf_path in pdf_files
            }
            
            for future in as_completed(futures):
                result = future.result()
                if result['success']:
                    jurisprudence_list.append(result)
        
        # Créer le fichier de sortie
        if output_format == 'csv':
            output_path, output_filename = create_jurisprudence_csv(jurisprudence_list, temp_folder)
        else:
            output_path, output_filename = create_jurisprudence_excel(jurisprudence_list, temp_folder)
        
        # Nettoyer les fichiers extraits
        import shutil
        shutil.rmtree(extract_dir, ignore_errors=True)
        
        return {
            'success': True,
            'output_path': output_path,
            'filename': output_filename,
            'total': len(pdf_files),
            'successful': len(jurisprudence_list),
            'failed': len(pdf_files) - len(jurisprudence_list)
        }
        
    except Exception as e:
        logger.error(f"Erreur extraction jurisprudence depuis ZIP: {str(e)}")
        return {
            'success': False,
            'error': str(e)
        }

def extract_jurisprudence_from_single_pdf(pdf_path, temp_folder, filename, output_format='excel'):
    """Extrait la jurisprudence depuis un seul PDF"""
    # Charger la clé API depuis .env (VPS) ou secrets Replit
    api_key = Config.OPENROUTER_API_KEY
    if not api_key:
        return {
            'success': False,
            'error': 'Clé API OpenRouter non configurée. Veuillez configurer OPENROUTER_API_KEY dans les secrets ou le fichier .env.'
        }
    
    try:
        result = process_single_pdf_jurisprudence(pdf_path, filename, api_key)
        
        if not result['success']:
            return {
                'success': False,
                'error': result.get('error', 'Erreur d\'extraction')
            }
        
        # Créer le fichier de sortie
        jurisprudence_list = [result]
        if output_format == 'csv':
            output_path, output_filename = create_jurisprudence_csv(jurisprudence_list, temp_folder)
        else:
            output_path, output_filename = create_jurisprudence_excel(jurisprudence_list, temp_folder)
        
        return {
            'success': True,
            'output_path': output_path,
            'filename': output_filename,
            'total': 1,
            'successful': 1,
            'failed': 0
        }
        
    except Exception as e:
        logger.error(f"Erreur extraction jurisprudence depuis PDF: {str(e)}")
        return {
            'success': False,
            'error': str(e)
        }
