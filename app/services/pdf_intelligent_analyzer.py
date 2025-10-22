"""
PdfTools
MOA Digital Agency LLC
Par : Aisance KALONJI
Mail : moa@myoneart.com
www.myoneart.com
"""

import os
import requests
import base64
import zipfile
import csv
import uuid
import logging
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
import tempfile
from app.utils.progress import progress_manager

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

OPENROUTER_API_KEY = os.environ.get('OPENROUTER_API_KEY')
OPENROUTER_API_URL = 'https://openrouter.ai/api/v1/chat/completions'

def analyze_pdf_with_openrouter(pdf_source, source_type='url', model='anthropic/claude-3.5-sonnet', filename='document.pdf'):
    """
    Analyse un PDF avec OpenRouter et extrait les données structurées
    
    Args:
        pdf_source: URL du PDF ou contenu base64
        source_type: 'url' ou 'base64'
        model: Modèle à utiliser pour l'analyse
        filename: Nom du fichier PDF
    
    Returns:
        dict: Données extraites du PDF
    """
    try:
        if not OPENROUTER_API_KEY:
            return {
                'success': False,
                'error': 'Clé API OpenRouter non configurée. Veuillez ajouter OPENROUTER_API_KEY dans les secrets.'
            }
        
        if source_type == 'url':
            pdf_content = {
                "type": "file",
                "file": {
                    "filename": filename,
                    "file_data": pdf_source
                }
            }
        else:  # base64
            pdf_content = {
                "type": "file",
                "file": {
                    "filename": filename,
                    "file_data": f"data:application/pdf;base64,{pdf_source}"
                }
            }
        
        headers = {
            'Authorization': f'Bearer {OPENROUTER_API_KEY}',
            'Content-Type': 'application/json'
        }
        
        payload = {
            "model": model,
            "messages": [
                {
                    "role": "user",
                    "content": [
                        pdf_content,
                        {
                            "type": "text",
                            "text": """Analysez EN DÉTAIL ce document PDF et extrayez TOUTES les données présentes, y compris le texte, les tableaux, les chiffres, et toutes les informations importantes.

IMPORTANT : 
- Lisez TOUTES les pages du document
- Extrayez TOUS les tableaux que vous trouvez
- Récupérez TOUS les champs de données structurés
- N'omettez aucune information importante
- Si le document contient des formulaires, extrayez tous les champs

Retournez les données au format JSON avec cette structure EXACTE :
{
    "metadata": {
        "titre": "titre complet du document",
        "date": "date si disponible",
        "auteur": "auteur si disponible",
        "entreprise": "entreprise/organisation si disponible",
        "type_document": "type de document (facture, contrat, rapport, etc.)",
        "numero_pages": "nombre de pages"
    },
    "tables": [
        {
            "nom": "titre/description de la table",
            "page": "numéro de page où se trouve la table",
            "colonnes": ["nom_colonne1", "nom_colonne2", "nom_colonne3"],
            "lignes": [
                ["valeur1", "valeur2", "valeur3"],
                ["valeur1", "valeur2", "valeur3"]
            ]
        }
    ],
    "informations_cles": {
        "section1": "contenu de la section 1",
        "section2": "contenu de la section 2",
        "montant_total": "montant si applicable",
        "reference": "numéro de référence si applicable"
    },
    "texte_complet": "résumé du contenu textuel principal du document"
}

Assurez-vous d'extraire TOUTES les tables et TOUTES les données structurées présentes dans le document."""
                        }
                    ]
                }
            ]
        }
        
        response = requests.post(OPENROUTER_API_URL, headers=headers, json=payload, timeout=300)
        response.raise_for_status()
        
        result = response.json()
        content = result['choices'][0]['message']['content']
        
        # Essayer d'extraire le JSON de la réponse
        import json
        import re
        
        # Chercher un bloc JSON dans la réponse
        json_match = re.search(r'\{.*\}', content, re.DOTALL)
        if json_match:
            try:
                extracted_data = json.loads(json_match.group())
                
                # S'assurer que les champs requis existent avec des valeurs par défaut
                if 'metadata' not in extracted_data:
                    extracted_data['metadata'] = {}
                
                metadata_defaults = {
                    'titre': '',
                    'date': '',
                    'auteur': '',
                    'entreprise': '',
                    'type_document': '',
                    'numero_pages': ''
                }
                for key, default_val in metadata_defaults.items():
                    if key not in extracted_data['metadata']:
                        extracted_data['metadata'][key] = default_val
                
                if 'tables' not in extracted_data:
                    extracted_data['tables'] = []
                if 'informations_cles' not in extracted_data:
                    extracted_data['informations_cles'] = {}
                if 'texte_complet' not in extracted_data:
                    extracted_data['texte_complet'] = content[:1000] if len(content) > 1000 else content
                
                return {
                    'success': True,
                    'data': extracted_data,
                    'raw_response': content
                }
            except json.JSONDecodeError as e:
                logger.warning(f"Erreur de parsing JSON: {str(e)}")
                # Retourner une structure par défaut avec le contenu brut
                return {
                    'success': True,
                    'data': {
                        'metadata': {
                            'titre': '',
                            'date': '',
                            'auteur': '',
                            'entreprise': '',
                            'type_document': '',
                            'numero_pages': ''
                        },
                        'tables': [],
                        'informations_cles': {},
                        'texte_complet': content[:1000] if len(content) > 1000 else content
                    },
                    'raw_response': content
                }
        else:
            # Pas de JSON trouvé, retourner une structure par défaut
            return {
                'success': True,
                'data': {
                    'metadata': {
                        'titre': '',
                        'date': '',
                        'auteur': '',
                        'entreprise': '',
                        'type_document': '',
                        'numero_pages': ''
                    },
                    'tables': [],
                    'informations_cles': {},
                    'texte_complet': content[:1000] if len(content) > 1000 else content
                },
                'raw_response': content
            }
            
    except Exception as e:
        logger.error(f"Erreur lors de l'analyse du PDF: {str(e)}")
        return {
            'success': False,
            'error': str(e)
        }

def process_pdf_from_url(url, idx):
    """Télécharge et analyse un PDF depuis une URL"""
    try:
        logger.info(f"Analyse du PDF {idx}: {url}")
        # Extraire le nom du fichier depuis l'URL
        filename = url.split('/')[-1] if '/' in url else f'document_{idx}.pdf'
        if not filename.endswith('.pdf'):
            filename = f'document_{idx}.pdf'
        
        result = analyze_pdf_with_openrouter(url, source_type='url', filename=filename)
        
        if result['success']:
            return {
                'success': True,
                'url': url,
                'index': idx,
                'data': result['data']
            }
        else:
            return {
                'success': False,
                'url': url,
                'index': idx,
                'error': result.get('error', 'Erreur inconnue')
            }
    except Exception as e:
        return {
            'success': False,
            'url': url,
            'index': idx,
            'error': str(e)
        }

def process_pdf_from_file(file_path, idx, filename):
    """Analyse un fichier PDF local"""
    try:
        logger.info(f"Analyse du PDF {idx}: {filename}")
        
        with open(file_path, 'rb') as f:
            pdf_content = base64.b64encode(f.read()).decode('utf-8')
        
        result = analyze_pdf_with_openrouter(pdf_content, source_type='base64', filename=filename)
        
        if result['success']:
            return {
                'success': True,
                'filename': filename,
                'index': idx,
                'data': result['data']
            }
        else:
            return {
                'success': False,
                'filename': filename,
                'index': idx,
                'error': result.get('error', 'Erreur inconnue')
            }
    except Exception as e:
        return {
            'success': False,
            'filename': filename,
            'index': idx,
            'error': str(e)
        }

def create_excel_from_analysis(analysis_results, temp_folder):
    """Crée un fichier Excel à partir des résultats d'analyse"""
    try:
        wb = Workbook()
        
        # Récupérer la feuille active et la renommer
        summary_sheet = wb.active
        summary_sheet.title = "Synthèse"  # type: ignore
        summary_sheet.append(['Index', 'Source', 'Statut', 'Titre', 'Type', 'Date', 'Auteur', 'Entreprise', 'Pages'])  # type: ignore
        
        for result in analysis_results:
            if result['success']:
                data = result.get('data', {})
                metadata = data.get('metadata', {})
                source = result.get('url') or result.get('filename', 'N/A')
                
                summary_sheet.append([  # type: ignore
                    result.get('index', ''),
                    source,
                    'Succès',
                    metadata.get('titre', ''),
                    metadata.get('type_document', ''),
                    metadata.get('date', ''),
                    metadata.get('auteur', ''),
                    metadata.get('entreprise', ''),
                    metadata.get('numero_pages', '')
                ])
                
                # Créer des feuilles pour chaque table trouvée
                tables = data.get('tables', [])
                for table_idx, table in enumerate(tables):
                    sheet_name = f"PDF_{result['index']}_Table_{table_idx + 1}"[:31]  # Limite Excel
                    table_sheet = wb.create_sheet(sheet_name)
                    
                    # Ajouter le nom de la table
                    table_sheet.append([table.get('nom', f'Table {table_idx + 1}')])
                    table_sheet.append([])  # Ligne vide
                    
                    # Ajouter les colonnes
                    columns = table.get('colonnes', [])
                    if columns:
                        table_sheet.append(columns)
                    
                    # Ajouter les lignes
                    rows = table.get('lignes', [])
                    for row in rows:
                        table_sheet.append(row)
                
                # Créer une feuille pour les informations clés
                info_keys = data.get('informations_cles', {})
                if info_keys:
                    info_sheet_name = f"PDF_{result['index']}_Infos"[:31]
                    info_sheet = wb.create_sheet(info_sheet_name)
                    info_sheet.append(['Clé', 'Valeur'])
                    for key, value in info_keys.items():
                        info_sheet.append([key, str(value)])
                
                # Créer une feuille pour le texte complet
                texte_complet = data.get('texte_complet', '')
                if texte_complet:
                    text_sheet_name = f"PDF_{result['index']}_Texte"[:31]
                    text_sheet = wb.create_sheet(text_sheet_name)
                    text_sheet.append(['Texte Complet'])
                    text_sheet.append([texte_complet])
            else:
                source = result.get('url') or result.get('filename', 'N/A')
                summary_sheet.append([  # type: ignore
                    result.get('index', ''),
                    source,
                    'Échec',
                    '',
                    '',
                    result.get('error', 'Erreur inconnue')
                ])
        
        # Sauvegarder le fichier Excel
        unique_id = str(uuid.uuid4())[:8]
        excel_filename = f'analyse_pdf_{unique_id}.xlsx'
        excel_path = os.path.join(temp_folder, excel_filename)
        
        wb.save(excel_path)
        
        return {
            'success': True,
            'excel_path': excel_path,
            'filename': excel_filename
        }
        
    except Exception as e:
        logger.error(f"Erreur lors de la création du fichier Excel: {str(e)}")
        return {
            'success': False,
            'error': str(e)
        }

def analyze_pdfs_from_csv(csv_content, temp_folder, max_workers=5, session_id=None):
    """Analyse des PDFs depuis un fichier CSV contenant des URLs"""
    try:
        # Lire le CSV
        urls = []
        csv_lines = csv_content.decode('utf-8').splitlines()
        csv_reader = csv.reader(csv_lines)
        
        for row in csv_reader:
            if row and row[0].strip():
                urls.append(row[0].strip())
        
        if not urls:
            return {'success': False, 'error': 'Aucune URL trouvée dans le fichier CSV'}
        
        if session_id:
            progress_manager.update(session_id,
                status='analyzing',
                total=len(urls),
                message=f'Analyse intelligente de {len(urls)} PDFs...'
            )
        
        logger.info(f"Analyse de {len(urls)} PDFs depuis le CSV")
        
        # Analyser les PDFs en parallèle
        results = []
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = [
                executor.submit(process_pdf_from_url, url, idx)
                for idx, url in enumerate(urls, 1)
            ]
            
            for future in as_completed(futures):
                result = future.result()
                results.append(result)
                
                if session_id:
                    progress_manager.update(session_id,
                        current=len(results),
                        successful=sum(1 for r in results if r['success']),
                        failed=sum(1 for r in results if not r['success'])
                    )
        
        # Créer le fichier Excel
        excel_result = create_excel_from_analysis(results, temp_folder)
        
        if excel_result['success']:
            return {
                'success': True,
                'excel_path': excel_result['excel_path'],
                'filename': excel_result['filename'],
                'total': len(urls),
                'successful': sum(1 for r in results if r['success']),
                'failed': sum(1 for r in results if not r['success'])
            }
        else:
            return excel_result
            
    except Exception as e:
        logger.error(f"Erreur lors de l'analyse depuis CSV: {str(e)}")
        return {'success': False, 'error': str(e)}

def analyze_pdfs_from_zip(zip_path, temp_folder, max_workers=5, session_id=None):
    """Analyse des PDFs depuis un fichier ZIP"""
    try:
        # Extraire le ZIP
        extract_dir = os.path.join(temp_folder, str(uuid.uuid4()))
        os.makedirs(extract_dir, exist_ok=True)
        
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(extract_dir)
        
        # Trouver tous les fichiers PDF
        pdf_files = []
        for root, dirs, files in os.walk(extract_dir):
            for file in files:
                if file.lower().endswith('.pdf'):
                    pdf_files.append(os.path.join(root, file))
        
        if not pdf_files:
            return {'success': False, 'error': 'Aucun fichier PDF trouvé dans le ZIP'}
        
        if session_id:
            progress_manager.update(session_id,
                status='analyzing',
                total=len(pdf_files),
                message=f'Analyse intelligente de {len(pdf_files)} PDFs depuis le ZIP...'
            )
        
        logger.info(f"Analyse de {len(pdf_files)} PDFs depuis le ZIP")
        
        # Analyser les PDFs en parallèle
        results = []
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = [
                executor.submit(process_pdf_from_file, pdf_file, idx, os.path.basename(pdf_file))
                for idx, pdf_file in enumerate(pdf_files, 1)
            ]
            
            for future in as_completed(futures):
                result = future.result()
                results.append(result)
                
                if session_id:
                    progress_manager.update(session_id,
                        current=len(results),
                        successful=sum(1 for r in results if r['success']),
                        failed=sum(1 for r in results if not r['success'])
                    )
        
        # Nettoyer les fichiers extraits
        for root, dirs, files in os.walk(extract_dir, topdown=False):
            for file in files:
                os.remove(os.path.join(root, file))
            for dir in dirs:
                os.rmdir(os.path.join(root, dir))
        os.rmdir(extract_dir)
        
        # Créer le fichier Excel
        excel_result = create_excel_from_analysis(results, temp_folder)
        
        if excel_result['success']:
            return {
                'success': True,
                'excel_path': excel_result['excel_path'],
                'filename': excel_result['filename'],
                'total': len(pdf_files),
                'successful': sum(1 for r in results if r['success']),
                'failed': sum(1 for r in results if not r['success'])
            }
        else:
            return excel_result
            
    except Exception as e:
        logger.error(f"Erreur lors de l'analyse depuis ZIP: {str(e)}")
        return {'success': False, 'error': str(e)}

def analyze_single_pdf(file_path, temp_folder, filename):
    """Analyse un seul fichier PDF"""
    try:
        logger.info(f"Analyse d'un seul PDF: {filename}")
        
        result = process_pdf_from_file(file_path, 1, filename)
        
        # Créer le fichier Excel
        excel_result = create_excel_from_analysis([result], temp_folder)
        
        if excel_result['success']:
            return {
                'success': True,
                'excel_path': excel_result['excel_path'],
                'filename': excel_result['filename'],
                'total': 1,
                'successful': 1 if result['success'] else 0,
                'failed': 0 if result['success'] else 1
            }
        else:
            return excel_result
            
    except Exception as e:
        logger.error(f"Erreur lors de l'analyse d'un seul PDF: {str(e)}")
        return {'success': False, 'error': str(e)}
