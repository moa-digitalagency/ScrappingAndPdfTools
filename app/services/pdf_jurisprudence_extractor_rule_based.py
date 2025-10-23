"""
PdfTools
MOA Digital Agency LLC
Extraction de jurisprudence basée sur des règles (SANS IA)
Plus rapide, plus robuste, sans timeout
"""

import os
import re
import logging
import uuid
import zipfile
import shutil
from typing import Dict, List, Optional
from pypdf import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from concurrent.futures import ThreadPoolExecutor, as_completed
import csv

logger = logging.getLogger(__name__)

class JurisprudenceExtractor:
    """Extracteur de jurisprudence basé sur des règles et regex"""
    
    @staticmethod
    def extract_text_from_pdf(pdf_path: str) -> Optional[str]:
        """Extrait le texte d'un fichier PDF"""
        try:
            reader = PdfReader(pdf_path)
            text = ""
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
            return text.strip()
        except Exception as e:
            logger.error(f"Erreur extraction texte de {pdf_path}: {str(e)}")
            return None
    
    @staticmethod
    def extract_field(text: str, pattern: str, default: str = "N/A") -> str:
        """Extrait un champ avec un pattern regex (supporte accents et variantes)"""
        match = re.search(pattern, text, re.MULTILINE | re.IGNORECASE | re.UNICODE)
        if match:
            value = match.group(1).strip()
            return value if value else default
        return default
    
    @staticmethod
    def extract_multiline_field(text: str, start_pattern: str, end_pattern: str, default: str = "N/A") -> str:
        """Extrait un champ multi-lignes entre deux patterns"""
        try:
            pattern = f"{start_pattern}(.*?){end_pattern}"
            match = re.search(pattern, text, re.DOTALL | re.IGNORECASE)
            if match:
                value = match.group(1).strip()
                value = re.sub(r'\s+', ' ', value)
                return value if value else default
            return default
        except Exception:
            return default
    
    @staticmethod
    def extract_jurisprudence_data(pdf_text: str, filename: str) -> Dict:
        """
        Extrait les données de jurisprudence avec des règles et regex
        SANS IA - Plus rapide et robuste
        """
        try:
            if not pdf_text or len(pdf_text) < 50:
                raise ValueError("Texte insuffisant pour extraction")
            
            data = {}
            data['fichier'] = filename
            
            # Extraction de la référence
            data['ref'] = JurisprudenceExtractor.extract_field(
                pdf_text, 
                r'Ref\s*:?\s*(\d+)',
                "N/A"
            )
            
            # Extraction du titre (première ligne généralement)
            title_match = re.search(r'Titre\s*:?\s*(.+?)(?:\n|Ref)', pdf_text, re.DOTALL | re.IGNORECASE)
            if title_match:
                data['titre'] = re.sub(r'\s+', ' ', title_match.group(1).strip())
            else:
                first_lines = pdf_text.split('\n')[:5]
                data['titre'] = ' '.join(first_lines).strip()[:200] if first_lines else "N/A"
            
            # Extraction de la juridiction
            data['juridiction'] = JurisprudenceExtractor.extract_field(
                pdf_text,
                r'Juridiction\s*:?\s*([^\n]+)',
                "N/A"
            )
            
            # Extraction pays/ville
            data['pays_ville'] = JurisprudenceExtractor.extract_field(
                pdf_text,
                r'Pays/Ville\s*:?\s*([^\n]+)',
                "N/A"
            )
            
            # Extraction numéro de décision (avec et sans accents)
            data['numero_decision'] = JurisprudenceExtractor.extract_field(
                pdf_text,
                r'N°?\s*de\s*d[eé]cision\s*:?\s*([^\n]+)',
                "N/A"
            )
            
            # Extraction date de décision (avec et sans accents)
            date_pattern = r'Date\s*de\s*d[eé]cision\s*:?\s*([^\n]+)'
            data['date_decision'] = JurisprudenceExtractor.extract_field(
                pdf_text,
                date_pattern,
                "N/A"
            )
            
            # Si pas trouvé, chercher format de date
            if data['date_decision'] == "N/A":
                date_format_match = re.search(r'(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})', pdf_text)
                if date_format_match:
                    data['date_decision'] = date_format_match.group(1)
            
            # Extraction numéro de dossier
            data['numero_dossier'] = JurisprudenceExtractor.extract_field(
                pdf_text,
                r'N°\s*de\s*dossier\s*:?\s*([^\n]+)',
                "N/A"
            )
            
            # Extraction type de décision
            data['type_decision'] = JurisprudenceExtractor.extract_field(
                pdf_text,
                r'Type\s*de\s*décision\s*:?\s*([^\n]+)',
                "N/A"
            )
            
            # Extraction chambre
            data['chambre'] = JurisprudenceExtractor.extract_field(
                pdf_text,
                r'Chambre\s*:?\s*([^\n]+)',
                "N/A"
            )
            
            # Extraction thème
            data['theme'] = JurisprudenceExtractor.extract_field(
                pdf_text,
                r'Thème\s*:?\s*([^\n]+)',
                "N/A"
            )
            
            # Extraction mots-clés
            mots_cles = JurisprudenceExtractor.extract_multiline_field(
                pdf_text,
                r'Mots\s*clés\s*:?',
                r'(?:Base\s*légale|Article|Résumé|$)',
                "N/A"
            )
            data['mots_cles'] = mots_cles
            
            # Extraction base légale - Articles
            articles = JurisprudenceExtractor.extract_multiline_field(
                pdf_text,
                r'Article\(s\)\s*:?',
                r'(?:Article\(s\)\s*:|Résumé|Source|$)',
                "N/A"
            )
            data['base_legale_articles'] = articles
            
            # Extraction base légale - Lois
            lois_pattern = r'Loi\s*n°\s*[\d\-]+.*?(?:\n|Article|Résumé)'
            lois_matches = re.findall(lois_pattern, pdf_text, re.IGNORECASE)
            if lois_matches:
                data['base_legale_lois'] = '; '.join([l.strip() for l in lois_matches])
            else:
                data['base_legale_lois'] = "N/A"
            
            # Extraction résumé en français
            resume_fr = JurisprudenceExtractor.extract_multiline_field(
                pdf_text,
                r'Résumé\s*en\s*français\s*:?',
                r'(?:Résumé\s*en\s*arabe|Texte\s*intégral|$)',
                "N/A"
            )
            data['resume_francais'] = resume_fr
            
            # Extraction résumé en arabe
            resume_ar = JurisprudenceExtractor.extract_multiline_field(
                pdf_text,
                r'Résumé\s*en\s*arabe\s*:?',
                r'(?:Texte\s*intégral|$)',
                "N/A"
            )
            data['resume_arabe'] = resume_ar
            
            # Extraction texte intégral (début)
            texte_integral = JurisprudenceExtractor.extract_multiline_field(
                pdf_text,
                r'Texte\s*intégral\s*:?',
                r'$',
                "N/A"
            )
            if texte_integral and texte_integral != "N/A":
                data['texte_integral_debut'] = texte_integral[:1000]
            else:
                data['texte_integral_debut'] = pdf_text[:1000] if len(pdf_text) > 1000 else pdf_text
            
            # Extraction source
            data['source'] = JurisprudenceExtractor.extract_field(
                pdf_text,
                r'Source\s*:?\s*([^\n]+)',
                "N/A"
            )
            
            return data
            
        except Exception as e:
            logger.error(f"Erreur extraction jurisprudence pour {filename}: {str(e)}")
            return {
                "fichier": filename,
                "ref": "N/A",
                "titre": f"Erreur: {str(e)}",
                "juridiction": "N/A",
                "pays_ville": "N/A",
                "numero_decision": "N/A",
                "date_decision": "N/A",
                "numero_dossier": "N/A",
                "type_decision": "N/A",
                "chambre": "N/A",
                "theme": "N/A",
                "mots_cles": "N/A",
                "base_legale_articles": "N/A",
                "base_legale_lois": "N/A",
                "resume_francais": "N/A",
                "resume_arabe": "N/A",
                "texte_integral_debut": "N/A",
                "source": "N/A"
            }
    
    @staticmethod
    def process_single_pdf(pdf_path: str, filename: str) -> Dict:
        """Traite un seul PDF de jurisprudence"""
        try:
            logger.info(f"Extraction jurisprudence: {filename}")
            
            text = JurisprudenceExtractor.extract_text_from_pdf(pdf_path)
            if not text or len(text) < 50:
                return {
                    'success': False,
                    'filename': filename,
                    'error': 'Texte insuffisant pour extraction'
                }
            
            jurisprudence_data = JurisprudenceExtractor.extract_jurisprudence_data(text, filename)
            return {
                'success': True,
                'filename': filename,
                'data': jurisprudence_data
            }
            
        except Exception as e:
            logger.error(f"Erreur traitement {filename}: {str(e)}")
            return {
                'success': False,
                'filename': filename,
                'error': str(e)
            }
    
    @staticmethod
    def create_excel(jurisprudence_list: List[Dict], temp_folder: str) -> tuple:
        """Crée un fichier Excel avec les données de jurisprudence"""
        try:
            unique_id = str(uuid.uuid4())[:8]
            excel_filename = f'jurisprudence_database_{unique_id}.xlsx'
            excel_path = os.path.join(temp_folder, excel_filename)
            
            wb = Workbook()
            ws = wb.active
            ws.title = 'Base Jurisprudence'
            
            # En-têtes
            headers = [
                'Fichier', 'Ref', 'Titre', 'Juridiction', 'Pays/Ville', 
                'N° Décision', 'Date Décision', 'N° Dossier', 'Type Décision',
                'Chambre', 'Thème', 'Mots-clés', 'Articles', 'Lois',
                'Résumé Français', 'Résumé Arabe', 'Extrait Texte Intégral', 'Source'
            ]
            ws.append(headers)
            
            # Style des en-têtes
            header_fill = PatternFill(start_color='1F4788', end_color='1F4788', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
            
            # Données
            for item in jurisprudence_list:
                data = item.get('data', {})
                ws.append([
                    data.get('fichier', 'N/A'),
                    data.get('ref', 'N/A'),
                    data.get('titre', 'N/A'),
                    data.get('juridiction', 'N/A'),
                    data.get('pays_ville', 'N/A'),
                    data.get('numero_decision', 'N/A'),
                    data.get('date_decision', 'N/A'),
                    data.get('numero_dossier', 'N/A'),
                    data.get('type_decision', 'N/A'),
                    data.get('chambre', 'N/A'),
                    data.get('theme', 'N/A'),
                    data.get('mots_cles', 'N/A'),
                    data.get('base_legale_articles', 'N/A'),
                    data.get('base_legale_lois', 'N/A'),
                    data.get('resume_francais', 'N/A'),
                    data.get('resume_arabe', 'N/A'),
                    data.get('texte_integral_debut', 'N/A'),
                    data.get('source', 'N/A')
                ])
            
            # Style des données
            for row in ws.iter_rows(min_row=2, max_row=len(jurisprudence_list) + 1):
                for cell in row:
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                    cell.border = border
            
            # Largeurs de colonnes
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 50
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 20
            ws.column_dimensions['I'].width = 15
            ws.column_dimensions['J'].width = 15
            ws.column_dimensions['K'].width = 30
            ws.column_dimensions['L'].width = 50
            ws.column_dimensions['M'].width = 30
            ws.column_dimensions['N'].width = 40
            ws.column_dimensions['O'].width = 70
            ws.column_dimensions['P'].width = 70
            ws.column_dimensions['Q'].width = 50
            ws.column_dimensions['R'].width = 20
            
            wb.save(excel_path)
            logger.info(f"Base de données Excel créée: {excel_filename}")
            
            return excel_path, excel_filename
            
        except Exception as e:
            logger.error(f"Erreur création Excel: {str(e)}")
            raise
    
    @staticmethod
    def create_csv(jurisprudence_list: List[Dict], temp_folder: str) -> tuple:
        """Crée un fichier CSV avec les données de jurisprudence"""
        try:
            unique_id = str(uuid.uuid4())[:8]
            csv_filename = f'jurisprudence_database_{unique_id}.csv'
            csv_path = os.path.join(temp_folder, csv_filename)
            
            headers = [
                'Fichier', 'Ref', 'Titre', 'Juridiction', 'Pays/Ville', 
                'N° Décision', 'Date Décision', 'N° Dossier', 'Type Décision',
                'Chambre', 'Thème', 'Mots-clés', 'Articles', 'Lois',
                'Résumé Français', 'Résumé Arabe', 'Extrait Texte', 'Source'
            ]
            
            with open(csv_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(headers)
                
                for item in jurisprudence_list:
                    data = item.get('data', {})
                    writer.writerow([
                        data.get('fichier', 'N/A'),
                        data.get('ref', 'N/A'),
                        data.get('titre', 'N/A'),
                        data.get('juridiction', 'N/A'),
                        data.get('pays_ville', 'N/A'),
                        data.get('numero_decision', 'N/A'),
                        data.get('date_decision', 'N/A'),
                        data.get('numero_dossier', 'N/A'),
                        data.get('type_decision', 'N/A'),
                        data.get('chambre', 'N/A'),
                        data.get('theme', 'N/A'),
                        data.get('mots_cles', 'N/A'),
                        data.get('base_legale_articles', 'N/A'),
                        data.get('base_legale_lois', 'N/A'),
                        data.get('resume_francais', 'N/A'),
                        data.get('resume_arabe', 'N/A'),
                        data.get('texte_integral_debut', 'N/A'),
                        data.get('source', 'N/A')
                    ])
            
            logger.info(f"Base de données CSV créée: {csv_filename}")
            return csv_path, csv_filename
            
        except Exception as e:
            logger.error(f"Erreur création CSV: {str(e)}")
            raise
    
    @staticmethod
    def extract_from_zip_both_formats(zip_path: str, temp_folder: str, max_workers: int = 4) -> Dict:
        """
        Extrait la jurisprudence et crée BOTH Excel et CSV en un seul passage
        Plus efficace - évite le double traitement
        """
        try:
            # Extraire le ZIP
            extract_dir = os.path.join(temp_folder, f'jurisprudence_extract_{uuid.uuid4().hex[:8]}')
            os.makedirs(extract_dir, exist_ok=True)
            
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                zip_ref.extractall(extract_dir)
            
            # Trouver tous les PDFs
            pdf_files = []
            for root, dirs, files in os.walk(extract_dir):
                for file in files:
                    if file.lower().endswith('.pdf'):
                        pdf_files.append(os.path.join(root, file))
            
            if not pdf_files:
                return {
                    'success': False,
                    'error': 'Aucun fichier PDF trouvé dans le ZIP'
                }
            
            logger.info(f"Traitement de {len(pdf_files)} fichiers PDF de jurisprudence")
            
            # Traiter les PDFs en parallèle - UN SEUL PASSAGE
            jurisprudence_list = []
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                futures = {
                    executor.submit(
                        JurisprudenceExtractor.process_single_pdf,
                        pdf_path,
                        os.path.basename(pdf_path)
                    ): pdf_path for pdf_path in pdf_files
                }
                
                for future in as_completed(futures):
                    result = future.result()
                    if result['success']:
                        jurisprudence_list.append(result)
            
            # Vérifier si au moins un PDF a été traité avec succès
            if len(jurisprudence_list) == 0:
                return {
                    'success': False,
                    'error': f'Aucun PDF n\'a pu être traité avec succès sur {len(pdf_files)} fichiers',
                    'total': len(pdf_files),
                    'successful': 0,
                    'failed': len(pdf_files)
                }
            
            # Créer les DEUX fichiers de sortie en un seul passage
            excel_path, excel_filename = JurisprudenceExtractor.create_excel(jurisprudence_list, temp_folder)
            csv_path, csv_filename = JurisprudenceExtractor.create_csv(jurisprudence_list, temp_folder)
            
            # Nettoyer les fichiers extraits
            shutil.rmtree(extract_dir, ignore_errors=True)
            
            return {
                'success': True,
                'excel_path': excel_path,
                'excel_filename': excel_filename,
                'csv_path': csv_path,
                'csv_filename': csv_filename,
                'total': len(pdf_files),
                'successful': len(jurisprudence_list),
                'failed': len(pdf_files) - len(jurisprudence_list)
            }
            
        except Exception as e:
            logger.error(f"Erreur extraction jurisprudence depuis ZIP: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }
    
    @staticmethod
    def extract_from_zip(zip_path: str, temp_folder: str, output_format: str = 'excel', max_workers: int = 4) -> Dict:
        """
        Extrait la jurisprudence depuis un ZIP de PDFs
        SANS IA - Plus rapide et robuste
        """
        try:
            # Extraire le ZIP
            extract_dir = os.path.join(temp_folder, f'jurisprudence_extract_{uuid.uuid4().hex[:8]}')
            os.makedirs(extract_dir, exist_ok=True)
            
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                zip_ref.extractall(extract_dir)
            
            # Trouver tous les PDFs
            pdf_files = []
            for root, dirs, files in os.walk(extract_dir):
                for file in files:
                    if file.lower().endswith('.pdf'):
                        pdf_files.append(os.path.join(root, file))
            
            if not pdf_files:
                return {
                    'success': False,
                    'error': 'Aucun fichier PDF trouvé dans le ZIP'
                }
            
            logger.info(f"Traitement de {len(pdf_files)} fichiers PDF de jurisprudence")
            
            # Traiter les PDFs en parallèle
            jurisprudence_list = []
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                futures = {
                    executor.submit(
                        JurisprudenceExtractor.process_single_pdf,
                        pdf_path,
                        os.path.basename(pdf_path)
                    ): pdf_path for pdf_path in pdf_files
                }
                
                for future in as_completed(futures):
                    result = future.result()
                    if result['success']:
                        jurisprudence_list.append(result)
            
            # Créer les fichiers de sortie
            if output_format == 'csv':
                output_path, output_filename = JurisprudenceExtractor.create_csv(jurisprudence_list, temp_folder)
            else:
                output_path, output_filename = JurisprudenceExtractor.create_excel(jurisprudence_list, temp_folder)
            
            # Nettoyer les fichiers extraits
            shutil.rmtree(extract_dir, ignore_errors=True)
            
            return {
                'success': True,
                'output_path': output_path,
                'filename': output_filename,
                'total': len(pdf_files),
                'successful': len(jurisprudence_list),
                'failed': len(pdf_files) - len(jurisprudence_list)
            }
            
        except Exception as e:
            logger.error(f"Erreur extraction jurisprudence depuis ZIP: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }
    
    @staticmethod
    def extract_from_single_pdf(pdf_path: str, temp_folder: str, filename: str, output_format: str = 'excel') -> Dict:
        """Extrait la jurisprudence depuis un seul PDF"""
        try:
            result = JurisprudenceExtractor.process_single_pdf(pdf_path, filename)
            
            if not result['success']:
                return {
                    'success': False,
                    'error': result.get('error', 'Erreur d\'extraction')
                }
            
            # Créer le fichier de sortie
            jurisprudence_list = [result]
            if output_format == 'csv':
                output_path, output_filename = JurisprudenceExtractor.create_csv(jurisprudence_list, temp_folder)
            else:
                output_path, output_filename = JurisprudenceExtractor.create_excel(jurisprudence_list, temp_folder)
            
            return {
                'success': True,
                'output_path': output_path,
                'filename': output_filename,
                'total': 1,
                'successful': 1,
                'failed': 0
            }
            
        except Exception as e:
            logger.error(f"Erreur extraction jurisprudence depuis PDF: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }
